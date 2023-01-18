from openpyxl import load_workbook


def validate_xlsx(path):
    try:
        wb = load_workbook(path)
        return True
    except:
        return False


def read_index_xlsx(path):
    # load excel ...
    wb = load_workbook(path, data_only=True)
    # check sheets
    sheet_idx = 0
    if len(wb.worksheets) > 1:
        sheet_idx = wb.sheetnames.index("Sync")
    sheet = wb.worksheets[sheet_idx]
    headers = dict((header, i + 1)
                   for i, header in enumerate(next(sheet.values)) if header is not None)

    # validate required column headers
    cols = ("Throw", "Camera", "Frame")
    if not all(header in headers for header in cols):
        raise Exception(
            "Index file '%s' sheet '%s' missing required header" % (path, sheet.title))

    row_num = 2
    rows = []
    empty_rows = 0
    while True:
        values = [sheet.cell(row_num, headers[col]).value for col in cols]
        print(values)
        if all(False if v is None else True for v in values):
            rows.append(values)
            empty_rows = 0
        else:
            empty_rows += 1
        row_num += 1
        if empty_rows > 2:
            break

    return rows, cols

